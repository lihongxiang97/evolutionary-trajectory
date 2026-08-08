#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import zipfile
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


FORMATS = {".pdf", ".png", ".svg", ".tiff"}


def doc_text(path: Path) -> tuple[str, int]:
    doc = Document(path)
    text = "\n".join(p.text for p in doc.paragraphs)
    return text, len(doc.inline_shapes)


def cited(text: str, pattern: str) -> set[int]:
    return {int(x) for x in re.findall(pattern, text, flags=re.I)}


def supplementary_figure_citations(text: str) -> set[int]:
    values: set[int] = set()
    for paragraph in text.splitlines():
        if re.search(r"Supplementary\s+Fig", paragraph, flags=re.I):
            values.update(int(x) for x in re.findall(r"\bS(\d+)\b", paragraph))
    return values


def formats(folder: Path, prefix: str) -> set[str]:
    return {p.suffix.lower() for p in folder.glob(prefix + ".*")}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=Path, required=True)
    ap.add_argument("--report", type=Path, required=True)
    args = ap.parse_args()
    root = args.root.resolve()
    main_doc = root / "Manuscript_T2T_hapA_hapB_3D_mechanisms.docx"
    supp_doc = root / "Supplementary_Information.docx"
    workbook = root / "Supplementary_Tables.xlsx"

    main_text, main_images = doc_text(main_doc)
    _, supp_images = doc_text(supp_doc)
    wb = load_workbook(workbook, read_only=True)

    main_figs = cited(main_text, r"\bFig\.\s*([1-8])(?:[A-F,]|\b)")
    supp_figs = supplementary_figure_citations(main_text)
    supp_tables = cited(main_text, r"Supplementary\s+Table(?:s)?\s+(\d+)")

    process_terms = re.findall(
        r"\b(reanalysis|reanalyzed|remapping|old assembly|previous assembly|risk audit)\b",
        main_text,
        flags=re.I,
    )
    lettered_tables = re.findall(r"Supplementary\s+Table\s+\d+[a-c]\b", main_text, flags=re.I)

    main_formats_ok = all(
        formats(root / "Figures_Main", next((p.stem for p in (root / "Figures_Main").glob(f"Figure{i}_*.pdf")), "")) == FORMATS
        for i in range(1, 9)
    )
    supp_formats_ok = all(
        formats(root / "Figures_Supplementary", next((p.stem for p in (root / "Figures_Supplementary").glob(f"Supplementary_Figure_S{i}_*.pdf")), "")) == FORMATS
        for i in range(1, 19)
    )

    checks = {
        "main DOCX ZIP integrity": zipfile.is_zipfile(main_doc),
        "supplement DOCX ZIP integrity": zipfile.is_zipfile(supp_doc),
        "main manuscript contains 8 figures": main_images == 8,
        "supplement contains 18 figures": supp_images == 18,
        "all main figures cited": main_figs == set(range(1, 9)),
        "all supplementary figures cited": supp_figs == set(range(1, 19)),
        "all supplementary tables cited": supp_tables == set(range(1, 22)),
        "supplement workbook contains Contents and Table 1-21": wb.sheetnames
        == ["Contents"] + [f"Table {i}" for i in range(1, 22)],
        "no letter-suffixed supplementary tables": not lettered_tables,
        "no workflow-history wording": not process_terms,
        "all main figures have PDF/SVG/PNG/TIFF": main_formats_ok,
        "all supplementary figures have PDF/SVG/PNG/TIFF": supp_formats_ok,
    }
    passed = all(checks.values())
    lines = [
        "# Submission package content audit",
        "",
        f"Overall content status: **{'PASS' if passed else 'REVIEW REQUIRED'}**",
        "",
    ]
    lines += [f"- {'PASS' if ok else 'FAIL'} - {name}" for name, ok in checks.items()]
    lines += [
        "",
        "## Administrative items requiring author input",
        "",
        "- Author names, affiliations, corresponding-author email.",
        "- Authors' contributions.",
        "- Funding statement.",
        "- Final public code archive/version and cover letter.",
    ]
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print("PASS" if passed else "REVIEW REQUIRED")
    if not passed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
